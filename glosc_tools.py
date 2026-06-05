from __future__ import annotations

import base64
import csv
import json
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import time
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Literal

from markitdown import MarkItDown


MatchMode = Literal["contains", "equals", "regex"]


@dataclass(frozen=True)
class SpawnResult:
    code: int
    output: str


class GloscTools:
    MAX_TEXT_FILE_SIZE = 1_000_000
    MAX_STORED_FILE_SUMMARIES = 200

    @staticmethod
    def iso_utc_now() -> str:
        return (
            datetime.now(timezone.utc)
            .isoformat(timespec="milliseconds")
            .replace("+00:00", "Z")
        )

    @staticmethod
    def escape_powershell_single_quoted(value: str) -> str:
        return "'" + value.replace("'", "''") + "'"

    @staticmethod
    def extract_json_from_output(output: str) -> str | None:
        trimmed = output.strip()
        if not trimmed:
            return None

        first_brace = trimmed.find("{")
        first_bracket = trimmed.find("[")
        starts = [n for n in (first_brace, first_bracket) if n >= 0]
        if not starts:
            return None
        start = min(starts)

        last_brace = trimmed.rfind("}")
        last_bracket = trimmed.rfind("]")
        ends = [n for n in (last_brace, last_bracket) if n >= 0]
        if not ends:
            return None
        end = max(ends)

        if end <= start:
            return None
        return trimmed[start : end + 1]

    @staticmethod
    def spawn_command(
        command: str, args: list[str], *, encoding: str = "utf-8"
    ) -> SpawnResult:
        proc = subprocess.run(
            [command, *args],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            check=False,
        )
        output = proc.stdout.decode(encoding, errors="replace")
        if proc.returncode != 0:
            raise RuntimeError(
                f"Process exited with code {proc.returncode}: {command} {' '.join(args)}\n{output}"
            )
        return SpawnResult(code=proc.returncode, output=output)

    @staticmethod
    def match_score(name: str, query: str) -> int:
        n = name.lower()
        q = query.lower()
        if n == q:
            return 0
        if n.startswith(q):
            return 1
        if q in n:
            return 2
        return 3

    @staticmethod
    def filter_apps(apps: list[dict[str, Any]], query: str, match_mode: MatchMode) -> list[dict[str, Any]]:
        q = query.strip()
        if not q:
            return apps

        if match_mode == "regex":
            try:
                pattern = re.compile(q, re.IGNORECASE)
            except re.error:
                return []
            return [a for a in apps if pattern.search(str(a.get("name", "")))]

        nq = q.lower()
        matches: list[dict[str, Any]] = []
        for app in apps:
            name = str(app.get("name", ""))
            n = name.lower()
            if match_mode == "equals":
                if n == nq:
                    matches.append(app)
            else:
                if nq in n:
                    matches.append(app)

        return sorted(matches, key=lambda a: GloscTools.match_score(str(a.get("name", "")), q))

    @staticmethod
    def list_installed_apps(options: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        if os.name != "nt":
            raise RuntimeError("当前仅支持 Windows：通过注册表获取已安装应用列表")

        options = options or {}
        query = (options.get("query") or "").strip() or None
        match_mode: MatchMode = options.get("matchMode") or "contains"
        limit = int(options.get("limit") or 200)

        ps_script = r"""
$ErrorActionPreference = 'SilentlyContinue'
$paths = @(
  'HKLM:\Software\Microsoft\Windows\CurrentVersion\Uninstall\*',
  'HKLM:\Software\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall\*',
  'HKCU:\Software\Microsoft\Windows\CurrentVersion\Uninstall\*'
)

$apps = foreach ($p in $paths) {
  Get-ItemProperty -Path $p | Where-Object { $_.DisplayName -and $_.DisplayName.Trim().Length -gt 0 } | ForEach-Object {
    [PSCustomObject]@{
      Name = $_.DisplayName
      Version = $_.DisplayVersion
      Publisher = $_.Publisher
      InstallLocation = $_.InstallLocation
      InstallSource = $_.InstallSource
      DisplayIcon = $_.DisplayIcon
      UninstallString = $_.UninstallString
      QuietUninstallString = $_.QuietUninstallString
      RegistryKey = $_.PSPath
    }
  }
}

$apps = $apps | Sort-Object Name, Version, Publisher -Unique

if ($null -ne $apps) {
  $apps | ConvertTo-Json -Depth 4
}
"""

        res = GloscTools.spawn_command(
            "powershell",
            ["-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
            encoding="utf-8",
        )

        json_text = GloscTools.extract_json_from_output(res.output) or "[]"
        try:
            parsed = json.loads(json_text)
        except json.JSONDecodeError as e:
            raise RuntimeError(f"解析应用列表失败: {e}")

        apps_raw = parsed if isinstance(parsed, list) else [parsed]

        normalized: list[dict[str, Any]] = []
        for a in apps_raw:
            if not isinstance(a, dict):
                continue
            name = a.get("Name")
            if not isinstance(name, str) or not name.strip():
                continue
            normalized.append(
                {
                    "name": str(name),
                    "version": str(a.get("Version")) if a.get("Version") else None,
                    "publisher": str(a.get("Publisher")) if a.get("Publisher") else None,
                    "installLocation": str(a.get("InstallLocation")) if a.get("InstallLocation") else None,
                    "installSource": str(a.get("InstallSource")) if a.get("InstallSource") else None,
                    "displayIcon": str(a.get("DisplayIcon")) if a.get("DisplayIcon") else None,
                    "uninstallString": str(a.get("UninstallString")) if a.get("UninstallString") else None,
                    "quietUninstallString": str(a.get("QuietUninstallString"))
                    if a.get("QuietUninstallString")
                    else None,
                    "registryKey": str(a.get("RegistryKey")) if a.get("RegistryKey") else None,
                }
            )

        normalized.sort(key=lambda x: str(x.get("name", "")).lower())
        filtered = GloscTools.filter_apps(normalized, query, match_mode) if query else normalized
        return filtered[: max(1, limit)]

    @staticmethod
    def _extract_path_from_command_like(value: str | None) -> str | None:
        if not value:
            return None
        v = value.strip()
        if not v:
            return None

        comma_index = v.find(",")
        if comma_index > 0:
            v = v[:comma_index]

        if (v.startswith('"') and v.endswith('"')) or (v.startswith("'") and v.endswith("'")):
            v = v[1:-1]

        m = re.match(r'^"([^"]+)"', v)
        if m and m.group(1):
            return m.group(1)

        first_token = re.split(r"\s+", v, maxsplit=1)[0]
        return first_token or None

    @staticmethod
    def infer_install_path_from_app(app: dict[str, Any]) -> dict[str, Any]:
        loc = (app.get("installLocation") or "").strip()
        if loc:
            return {"installPath": loc, "source": "InstallLocation", "raw": loc}

        from_display_icon = GloscTools._extract_path_from_command_like(app.get("displayIcon"))
        if from_display_icon:
            return {
                "installPath": str(Path(from_display_icon).parent),
                "source": "DisplayIcon",
                "raw": app.get("displayIcon"),
            }

        from_uninstall = GloscTools._extract_path_from_command_like(
            app.get("quietUninstallString") or app.get("uninstallString")
        )
        if from_uninstall:
            return {
                "installPath": str(Path(from_uninstall).parent),
                "source": "UninstallString",
                "raw": app.get("quietUninstallString") or app.get("uninstallString"),
            }

        return {}

    @staticmethod
    def get_app_install_path(options: dict[str, Any]) -> dict[str, Any]:
        query = str(options.get("name") or "").strip()
        if not query:
            raise RuntimeError("name 不能为空")

        match_mode: MatchMode = options.get("matchMode") or "contains"
        all_matches = bool(options.get("allMatches") or False)
        limit = int(options.get("limit") or 50)

        apps = GloscTools.list_installed_apps({"query": query, "matchMode": match_mode, "limit": limit})

        candidates: list[dict[str, Any]] = []
        for a in apps:
            inferred = GloscTools.infer_install_path_from_app(a)
            candidates.append(
                {
                    "name": a.get("name"),
                    "version": a.get("version"),
                    "publisher": a.get("publisher"),
                    "installLocation": a.get("installLocation"),
                    "inferredInstallPath": inferred.get("installPath"),
                    "inferredSource": inferred.get("source"),
                }
            )

        best = candidates[0] if candidates else None
        return {
            "query": query,
            "matchMode": match_mode,
            "result": (
                {
                    "name": best.get("name"),
                    "installPath": best.get("installLocation") or best.get("inferredInstallPath"),
                    "source": "InstallLocation" if best.get("installLocation") else best.get("inferredSource"),
                    "version": best.get("version"),
                    "publisher": best.get("publisher"),
                }
                if (best and not all_matches)
                else None
            ),
            "candidates": candidates if all_matches else candidates[:10],
        }

    @staticmethod
    def open_reference(options: dict[str, Any]) -> dict[str, Any]:
        target = str(options.get("target") or "").strip()
        if not target:
            raise RuntimeError("target 不能为空")

        args = options.get("args") or []
        wait = bool(options.get("wait") or False)

        if os.name == "nt":
            file_path = GloscTools.escape_powershell_single_quoted(target)
            arg_list = ""
            if args:
                arg_list = " -ArgumentList @(" + ", ".join(
                    GloscTools.escape_powershell_single_quoted(str(a)) for a in args
                ) + ")"
            wait_flag = " -Wait" if wait else ""
            script = f"Start-Process -FilePath {file_path}{arg_list}{wait_flag}"

            GloscTools.spawn_command(
                "powershell",
                ["-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
                encoding="utf-8",
            )
            return {"ok": True, "target": target}

        if sys.platform == "darwin":
            GloscTools.spawn_command("open", [target], encoding="utf-8")
            return {"ok": True, "target": target}

        GloscTools.spawn_command("xdg-open", [target], encoding="utf-8")
        return {"ok": True, "target": target}

    @staticmethod
    async def usebrowser(
        url: str,
        use_browser: bool = False,
        content_type: Literal["text", "html", "json"] = "text",
    ) -> str:
        import httpx

        try:
            if use_browser:
                try:
                    from playwright.async_api import async_playwright  # type: ignore
                except Exception:
                    return "获取网页内容失败: useBrowser=true 需要安装 playwright"

                async with async_playwright() as p:
                    browser = await p.chromium.launch()
                    page = await browser.new_page()
                    await page.goto(url, wait_until="networkidle")
                    html = await page.content()
                    await browser.close()

                if content_type == "html":
                    return html

                text = GloscTools._html_to_text(html)
                if content_type == "text":
                    return text

                try:
                    parsed = json.loads(text)
                    return json.dumps(parsed, ensure_ascii=False, indent=2)
                except Exception:
                    return text

            async with httpx.AsyncClient(follow_redirects=True, timeout=30) as client:
                resp = await client.get(url)
                resp.raise_for_status()

                if content_type == "json":
                    try:
                        parsed = resp.json()
                        return json.dumps(parsed, ensure_ascii=False, indent=2)
                    except Exception:
                        return resp.text

                return resp.text

        except Exception as e:
            return f"获取网页内容失败: {e}"

    @staticmethod
    def _html_to_text(html: str) -> str:
        try:
            from bs4 import BeautifulSoup  # type: ignore

            soup = BeautifulSoup(html, "html.parser")
            return soup.get_text(" ", strip=True)
        except Exception:
            # 简易兜底：去标签 + 压缩空白
            text = re.sub(r"<[^>]+>", " ", html)
            return re.sub(r"\s+", " ", text).strip()

    @staticmethod
    def _normalize_rel_path(value: str) -> str:
        normalized = str(value or "").replace("\\", "/")
        normalized = re.sub(r"^/+", "", normalized)
        normalized = re.sub(r"/+", "/", normalized)
        normalized = re.sub(r"(^|/)\.\.(?=/|$)", "", normalized)
        normalized = re.sub(r"^\./", "", normalized).strip()
        return "" if normalized == "." else normalized

    @staticmethod
    def _dirname_like(value: str) -> str:
        normalized = GloscTools._normalize_rel_path(value)
        index = normalized.rfind("/")
        return "" if index == -1 else normalized[:index]

    @staticmethod
    def _basename_like(value: str) -> str:
        normalized = GloscTools._normalize_rel_path(value)
        index = normalized.rfind("/")
        return normalized if index == -1 else normalized[index + 1 :]

    @staticmethod
    def _extname_like(value: str) -> str:
        base = GloscTools._basename_like(value)
        index = base.rfind(".")
        return "" if index == -1 else base[index:].lower()

    @staticmethod
    def _to_rel_path(root: Path, full_path: Path) -> str:
        try:
            return GloscTools._normalize_rel_path(str(full_path.relative_to(root)))
        except ValueError:
            return GloscTools._normalize_rel_path(str(full_path))

    @staticmethod
    def _should_read_skill_text(path: str, size: int) -> bool:
        if size > GloscTools.MAX_TEXT_FILE_SIZE:
            return False
        return GloscTools._extname_like(path) in {
            ".md",
            ".markdown",
            ".json",
            ".jsonc",
            ".yaml",
            ".yml",
            ".txt",
        }

    @staticmethod
    def _read_directory_bundle(root_path: str) -> list[dict[str, Any]]:
        root = Path(root_path)
        files: list[dict[str, Any]] = []

        for current_root, _, filenames in os.walk(root):
            current = Path(current_root)
            for name in filenames:
                file_path = current / name
                try:
                    info = file_path.stat()
                except OSError:
                    continue

                rel_path = GloscTools._to_rel_path(root, file_path)
                if not rel_path:
                    continue

                item: dict[str, Any] = {
                    "path": rel_path,
                    "size": int(info.st_size or 0),
                }

                if GloscTools._should_read_skill_text(rel_path, int(info.st_size or 0)):
                    try:
                        item["text"] = GloscTools._decode_text_bytes(file_path.read_bytes())
                    except OSError:
                        item["text"] = ""

                files.append(item)

        return files

    @staticmethod
    def _strip_inline_comment(value: str) -> str:
        in_single = False
        in_double = False
        escaped = False

        for index, char in enumerate(value):
            if escaped:
                escaped = False
                continue
            if char == "\\" and in_double:
                escaped = True
                continue
            if char == "'" and not in_double:
                in_single = not in_single
                continue
            if char == '"' and not in_single:
                in_double = not in_double
                continue
            if char == "#" and not in_single and not in_double:
                if index == 0 or value[index - 1].isspace():
                    return value[:index].rstrip()

        return value.strip()

    @staticmethod
    def _parse_frontmatter_scalar(value: str) -> Any:
        raw = GloscTools._strip_inline_comment(value).strip()
        if not raw:
            return ""

        if (raw.startswith('"') and raw.endswith('"')) or (raw.startswith("'") and raw.endswith("'")):
            return raw[1:-1]

        if raw.startswith("[") and raw.endswith("]"):
            try:
                parsed = json.loads(raw.replace("'", '"'))
                if isinstance(parsed, list):
                    return parsed
            except Exception:
                pass

        if raw.startswith("{") and raw.endswith("}"):
            try:
                parsed = json.loads(raw.replace("'", '"'))
                if isinstance(parsed, dict):
                    return parsed
            except Exception:
                pass

        return raw

    @staticmethod
    def _parse_skill_frontmatter(raw: str) -> dict[str, Any]:
        frontmatter: dict[str, Any] = {}
        current_key: str | None = None

        for line in raw.splitlines():
            if not line.strip() or line.lstrip().startswith("#"):
                continue

            list_match = re.match(r"^\s*-\s+(.+?)\s*$", line)
            if list_match and current_key:
                current = frontmatter.get(current_key)
                if not isinstance(current, list):
                    current = []
                    frontmatter[current_key] = current
                current.append(GloscTools._parse_frontmatter_scalar(list_match.group(1)))
                continue

            nested_match = re.match(r"^\s+([A-Za-z0-9_.-]+):\s*(.*?)\s*$", line)
            if nested_match and current_key:
                current = frontmatter.get(current_key)
                if not isinstance(current, dict):
                    current = {}
                    frontmatter[current_key] = current
                current[nested_match.group(1)] = GloscTools._parse_frontmatter_scalar(nested_match.group(2))
                continue

            match = re.match(r"^([A-Za-z0-9_.-]+):\s*(.*?)\s*$", line)
            if not match:
                continue

            key = match.group(1)
            value = match.group(2)
            current_key = key

            if value.strip():
                frontmatter[key] = GloscTools._parse_frontmatter_scalar(value)
            elif key == "metadata":
                frontmatter[key] = {}
            else:
                frontmatter[key] = []

        return frontmatter

    @staticmethod
    def _parse_skill_markdown(markdown: str) -> dict[str, Any]:
        text = str(markdown or "")
        matched = re.match(r"^---\s*\r?\n([\s\S]*?)\r?\n---\s*(?:\r?\n)?([\s\S]*)$", text)
        if not matched:
            return {"frontmatter": {}, "body": text.strip()}

        return {
            "frontmatter": GloscTools._parse_skill_frontmatter(matched.group(1) or ""),
            "body": (matched.group(2) or "").strip(),
        }

    @staticmethod
    def _slugify_skill_name(value: str) -> str:
        normalized = re.sub(r"[^a-z0-9-]+", "-", str(value or "").strip().lower())
        normalized = re.sub(r"-{2,}", "-", normalized).strip("-")
        return normalized or "imported-skill"

    @staticmethod
    def _unique_strings(values: list[Any]) -> list[str]:
        out: list[str] = []
        seen: set[str] = set()
        for value in values:
            text = str(value or "").strip()
            if not text or text in seen:
                continue
            seen.add(text)
            out.append(text)
        return out

    @staticmethod
    def _split_allowed_tools(value: Any) -> list[str]:
        if isinstance(value, list):
            return GloscTools._unique_strings([str(item).strip() for item in value])

        raw = str(value or "").strip()
        if not raw:
            return []
        return GloscTools._unique_strings(re.split(r"[\s,]+", raw))

    @staticmethod
    def _to_string_record(value: Any) -> dict[str, str]:
        if not isinstance(value, dict):
            return {}
        out: dict[str, str] = {}
        for key, raw_value in value.items():
            if raw_value is None:
                continue
            out[str(key)] = raw_value if isinstance(raw_value, str) else str(raw_value)
        return out

    @staticmethod
    def _parse_json_text(text: str) -> Any | None:
        try:
            without_comments = re.sub(r"//.*?$|/\*[\s\S]*?\*/", "", text, flags=re.MULTILINE)
            return json.loads(without_comments)
        except Exception:
            return None

    @staticmethod
    def _looks_like_mcp_config(file: dict[str, Any]) -> bool:
        text = str(file.get("text") or "")
        extension = GloscTools._extname_like(str(file.get("path") or ""))
        if extension not in {".json", ".jsonc", ".yaml", ".yml"}:
            return False

        if extension in {".json", ".jsonc"}:
            parsed = GloscTools._parse_json_text(text)
            if isinstance(parsed, dict):
                if "mcpServers" in parsed or "servers" in parsed:
                    return True
                return ("command" in parsed or "url" in parsed) and any(
                    key in parsed for key in ("args", "env", "headers", "type")
                )

        return bool(
            re.search(r"(^|\n)\s*(mcpServers|servers)\s*:", text)
            or re.search(r"(^|\n)\s*(command|url)\s*:", text)
        )

    @staticmethod
    def _parse_openclaw_plugin_meta(file: dict[str, Any]) -> dict[str, Any] | None:
        if GloscTools._basename_like(str(file.get("path") or "")) != "openclaw.plugin.json":
            return None

        parsed = GloscTools._parse_json_text(str(file.get("text") or ""))
        if not isinstance(parsed, dict):
            return None

        capabilities = parsed.get("capabilities") if isinstance(parsed.get("capabilities"), dict) else {}
        capability_tags = capabilities.get("capabilityTags") if isinstance(capabilities, dict) else []
        tool_names = capabilities.get("toolNames") if isinstance(capabilities, dict) else []
        bundled_skills = capabilities.get("bundledSkills") if isinstance(capabilities, dict) else []

        return {
            "kind": "openclaw-plugin",
            **({"packageName": str(parsed.get("name")).strip()} if parsed.get("name") else {}),
            **({"displayName": str(parsed.get("displayName")).strip()} if parsed.get("displayName") else {}),
            **(
                {"summary": str(parsed.get("summary") or parsed.get("description")).strip()}
                if parsed.get("summary") or parsed.get("description")
                else {}
            ),
            **({"runtimeId": str(parsed.get("runtimeId")).strip()} if parsed.get("runtimeId") else {}),
            "capabilityTags": GloscTools._unique_strings(capability_tags if isinstance(capability_tags, list) else []),
            "toolNames": GloscTools._unique_strings(tool_names if isinstance(tool_names, list) else []),
            "bundledSkillNames": GloscTools._unique_strings(bundled_skills if isinstance(bundled_skills, list) else []),
        }

    @staticmethod
    def _classify_skill_file(rel_path: str) -> str:
        normalized = GloscTools._normalize_rel_path(rel_path)
        if normalized.startswith("scripts/"):
            return "script"
        if normalized.startswith("references/"):
            return "reference"
        if normalized.startswith("assets/"):
            return "asset"
        return "other"

    @staticmethod
    def _collect_skill_files(root_dir: str, files: list[dict[str, Any]]) -> list[dict[str, Any]]:
        prefix = f"{GloscTools._normalize_rel_path(root_dir)}/" if root_dir else ""
        out: list[dict[str, Any]] = []

        for file in files:
            normalized = GloscTools._normalize_rel_path(str(file.get("path") or ""))
            if not normalized:
                continue
            if not prefix and normalized == "SKILL.md":
                continue
            if prefix:
                if not normalized.startswith(prefix) or normalized == f"{prefix}SKILL.md":
                    continue
                relative_path = normalized[len(prefix) :]
            else:
                relative_path = normalized

            out.append(
                {
                    "path": relative_path,
                    "size": int(file.get("size") or 0),
                    "kind": GloscTools._classify_skill_file(relative_path),
                }
            )

            if len(out) >= GloscTools.MAX_STORED_FILE_SUMMARIES:
                break

        return out

    @staticmethod
    def _import_skill_directory_source(directory_path: str) -> dict[str, Any]:
        source_path = str(directory_path or "").strip()
        source = {
            "kind": "directory",
            "original": source_path,
            "canonical": source_path,
            "label": Path(source_path).name or source_path,
        }

        files = GloscTools._read_directory_bundle(source_path)
        for file in files:
            file["path"] = GloscTools._normalize_rel_path(str(file.get("path") or ""))

        plugin_file = next(
            (file for file in files if GloscTools._basename_like(str(file.get("path") or "")) == "openclaw.plugin.json"),
            None,
        )
        package_meta = GloscTools._parse_openclaw_plugin_meta(plugin_file) if plugin_file else None
        bundled_mcp_count = sum(1 for file in files if GloscTools._looks_like_mcp_config(file))
        skill_files = [
            file
            for file in files
            if GloscTools._basename_like(str(file.get("path") or "")).upper() == "SKILL.MD"
        ]

        imported_at = int(time.time() * 1000)
        skills: list[dict[str, Any]] = []

        for index, skill_file in enumerate(skill_files):
            raw_markdown = str(skill_file.get("text") or "")
            parsed = GloscTools._parse_skill_markdown(raw_markdown)
            frontmatter = parsed["frontmatter"] if isinstance(parsed.get("frontmatter"), dict) else {}
            skill_root = GloscTools._dirname_like(str(skill_file.get("path") or ""))
            folder_name = GloscTools._basename_like(skill_root)
            fallback_name = (
                (package_meta or {}).get("displayName")
                or (package_meta or {}).get("packageName")
                or folder_name
                or f"imported-skill-{index + 1}"
            )
            name = str(frontmatter.get("name") or fallback_name).strip() or str(fallback_name)
            description = (
                str(frontmatter.get("description") or (package_meta or {}).get("summary") or f"{name} 导入的兼容技能").strip()
                or f"{name} 导入的兼容技能"
            )
            slug = GloscTools._slugify_skill_name(name)
            skill_warnings: list[str] = []

            if not frontmatter.get("name"):
                skill_warnings.append("缺少标准 name frontmatter，已使用兼容回退名称。")
            if not frontmatter.get("description"):
                skill_warnings.append("缺少标准 description frontmatter，已使用兼容回退描述。")

            ecosystem_tags = GloscTools._unique_strings(
                [
                    "agent-skills",
                    (package_meta or {}).get("kind"),
                    "bundled-mcp" if bundled_mcp_count > 0 else "",
                ]
            )

            skill: dict[str, Any] = {
                "id": str(uuid.uuid4()),
                "dedupeKey": f"{source['canonical']}::{slug}",
                "slug": slug,
                "name": name,
                "description": description,
                "rawMarkdown": raw_markdown,
                "instructions": str(parsed.get("body") or ""),
                "compatibility": str(frontmatter.get("compatibility") or "").strip(),
                "license": str(frontmatter.get("license") or "").strip(),
                "allowedTools": GloscTools._split_allowed_tools(frontmatter.get("allowed-tools")),
                "metadata": GloscTools._to_string_record(frontmatter.get("metadata")),
                "enabled": True,
                "importedAt": imported_at,
                "updatedAt": imported_at,
                "source": source,
                "ecosystemTags": ecosystem_tags,
                "warnings": skill_warnings,
                "files": GloscTools._collect_skill_files(skill_root, files),
                "bundledMcpCount": bundled_mcp_count,
            }
            if package_meta:
                skill["packageMeta"] = package_meta

            skills.append(skill)

        warnings: list[str] = []
        if not skill_files and package_meta and "plugin" in str(package_meta.get("kind") or ""):
            warnings.append("已识别 OpenClaw/ClawHub 插件元数据，但包内未发现可导入的 SKILL.md。")

        if not skills and bundled_mcp_count == 0:
            raise RuntimeError("未在导入内容中发现可兼容的 Skill 或 MCP 配置")

        return {"skills": skills, "warnings": warnings}

    @staticmethod
    def _read_skills_directory(directory_path: str) -> dict[str, Any]:
        clean_root = str(directory_path or "").strip()
        root = Path(clean_root)
        if not clean_root or not root.exists():
            return {"skills": [], "warnings": ["目录不存在或无法访问"], "scannedCandidates": []}
        if not root.is_dir():
            return {"skills": [], "warnings": ["路径不是目录"], "scannedCandidates": []}

        candidates: list[str] = []
        root_skill_detected = False
        try:
            entries = list(root.iterdir())
        except OSError as e:
            return {"skills": [], "warnings": [f"目录无法读取：{e}"], "scannedCandidates": []}

        for entry in entries:
            if entry.is_dir():
                candidates.append(str(entry))
                continue
            if entry.name.lower() == "skill.md":
                root_skill_detected = True

        if root_skill_detected:
            candidates.insert(0, clean_root)

        skills: list[dict[str, Any]] = []
        warnings: list[str] = []

        for candidate in candidates:
            try:
                imported = GloscTools._import_skill_directory_source(candidate)
                for skill in imported.get("skills", []):
                    metadata = dict(skill.get("metadata") or {})
                    metadata["skillDirectory.root"] = clean_root
                    skill["metadata"] = metadata
                    skill["ecosystemTags"] = GloscTools._unique_strings(
                        [*list(skill.get("ecosystemTags") or []), "managed-directory-skill"]
                    )
                    skill["warnings"] = GloscTools._unique_strings(list(skill.get("warnings") or []))
                    skills.append(skill)

                warnings.extend(
                    [f"{Path(candidate).name or candidate}：{item}" for item in imported.get("warnings", [])]
                )
            except Exception as e:
                warnings.append(f"{Path(candidate).name or candidate}：{e}")

        deduped: dict[str, dict[str, Any]] = {}
        for skill in skills:
            deduped[str(skill.get("dedupeKey") or skill.get("id") or uuid.uuid4())] = skill

        return {
            "skills": list(deduped.values()),
            "warnings": GloscTools._unique_strings(warnings),
            "scannedCandidates": candidates,
        }

    @staticmethod
    def _normalize_lock_slug(value: str) -> str:
        return GloscTools._slugify_skill_name(value)

    @staticmethod
    def _extract_clawhub_lock_entries(raw: Any, fallback_key: str | None = None) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []

        def append(entry: dict[str, Any]) -> None:
            slug = GloscTools._normalize_lock_slug(str(entry.get("slug") or ""))
            if not slug:
                return
            existing = next((item for item in out if item.get("slug") == slug), None)
            if existing:
                if not existing.get("version") and entry.get("version"):
                    existing["version"] = entry.get("version")
                if not existing.get("name") and entry.get("name"):
                    existing["name"] = entry.get("name")
                return
            next_entry: dict[str, Any] = {"slug": slug}
            if entry.get("version"):
                next_entry["version"] = entry.get("version")
            if entry.get("name"):
                next_entry["name"] = entry.get("name")
            out.append(next_entry)

        def walk(value: Any, key_hint: str | None = None, depth: int = 0) -> None:
            if depth > 5 or value is None:
                return
            if isinstance(value, list):
                for item in value:
                    walk(item, None, depth + 1)
                return
            if not isinstance(value, dict):
                return

            slug_candidate = str(
                value.get("slug") or value.get("skillSlug") or value.get("name") or key_hint or ""
            ).strip()
            version_candidate = str(value.get("version") or value.get("installedVersion") or "").strip()
            name_candidate = str(value.get("displayName") or value.get("title") or value.get("name") or "").strip()

            if slug_candidate and (version_candidate or name_candidate or value.get("slug")):
                append(
                    {
                        "slug": slug_candidate,
                        "version": version_candidate or None,
                        "name": name_candidate or None,
                    }
                )

            for key, nested in value.items():
                walk(nested, str(key), depth + 1)

        walk(raw, fallback_key)
        return out

    @staticmethod
    def _read_clawhub_lock(workspace_root: str) -> dict[str, Any]:
        lock_path = Path(workspace_root) / ".clawhub" / "lock.json"
        if not lock_path.exists():
            return {"lockPath": None, "entries": []}

        try:
            parsed = GloscTools._parse_json_text(GloscTools._decode_text_bytes(lock_path.read_bytes()))
            return {"lockPath": str(lock_path), "entries": GloscTools._extract_clawhub_lock_entries(parsed)}
        except Exception:
            return {"lockPath": str(lock_path), "entries": []}

    @staticmethod
    def _read_workspace_installed_skills(workspace_root: str) -> dict[str, Any]:
        clean_root = str(workspace_root or "").strip()
        if not clean_root:
            return {"skills": [], "warnings": [], "skillsDir": None, "lockPath": None, "lockEntries": []}

        skills_dir = Path(clean_root) / "skills"
        lock = GloscTools._read_clawhub_lock(clean_root)
        if not skills_dir.exists():
            return {
                "skills": [],
                "warnings": ["检测到 .clawhub/lock.json，但当前工作区不存在 skills/ 目录。"] if lock.get("lockPath") else [],
                "skillsDir": None,
                "lockPath": lock.get("lockPath"),
                "lockEntries": lock.get("entries") or [],
            }

        result = GloscTools._read_skills_directory(str(skills_dir))
        skills: list[dict[str, Any]] = []
        lock_entries = list(lock.get("entries") or [])

        for skill in result.get("skills", []):
            source_label = Path(str(skill.get("source", {}).get("canonical") or skill.get("source", {}).get("original") or "")).name
            source_slug = GloscTools._normalize_lock_slug(source_label or str(skill.get("slug") or ""))
            lock_entry = next((entry for entry in lock_entries if entry.get("slug") == source_slug), None)
            if not lock_entry:
                lock_entry = next((entry for entry in lock_entries if entry.get("slug") == skill.get("slug")), None)

            metadata = dict(skill.get("metadata") or {})
            metadata["clawhub.workspaceRoot"] = clean_root
            if lock_entry and lock_entry.get("version"):
                metadata["clawhub.version"] = str(lock_entry.get("version"))
            skill["metadata"] = metadata
            skill["ecosystemTags"] = GloscTools._unique_strings(
                [*list(skill.get("ecosystemTags") or []), "workspace-skill", "clawhub-installed"]
            )
            skill["warnings"] = GloscTools._unique_strings(list(skill.get("warnings") or []))
            skills.append(skill)

        deduped: dict[str, dict[str, Any]] = {}
        for skill in skills:
            deduped[str(skill.get("dedupeKey") or skill.get("id") or uuid.uuid4())] = skill

        return {
            "skills": list(deduped.values()),
            "warnings": result.get("warnings") or [],
            "skillsDir": str(skills_dir),
            "lockPath": lock.get("lockPath"),
            "lockEntries": lock_entries,
        }

    @staticmethod
    def read_skills(options: dict[str, Any]) -> dict[str, Any]:
        mode = str(options.get("mode") or "directory").strip() or "directory"
        path = str(options.get("path") or "").strip()
        if mode == "workspace":
            return GloscTools._read_workspace_installed_skills(path)
        if mode == "directory":
            return GloscTools._read_skills_directory(path)
        raise RuntimeError("mode 仅支持 directory 或 workspace")

    @staticmethod
    def parse_7z_output(output: str) -> list[dict[str, Any]]:
        lines = output.splitlines()
        file_list: list[dict[str, Any]] = []
        in_file_list = False

        for line in lines:
            if "Date" in line and "Time" in line and "Attr" in line:
                in_file_list = True
                continue

            if in_file_list and line.strip() and ("---" not in line) and ("files" not in line):
                parts = re.split(r"\s+", line.strip())
                if len(parts) >= 6:
                    attr = parts[2]
                    try:
                        size = int(parts[3])
                    except Exception:
                        size = 0
                    name = " ".join(parts[5:])
                    is_dir = ("D" in attr) or ("d" in attr)
                    file_list.append({"name": name, "size": size, "isDirectory": is_dir})

        return file_list

    @staticmethod
    def _detect_compound_ext(file_path: Path) -> str:
        lower = file_path.name.lower()
        for ext in [".tar.gz", ".tar.bz2"]:
            if lower.endswith(ext):
                return ext
        return file_path.suffix.lower()

    @staticmethod
    def process_file(file_path: str) -> str:
        p = Path(file_path)
        ext = GloscTools._detect_compound_ext(p)

        data = p.read_bytes()

        if ext in {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}:
            b64 = base64.b64encode(data).decode("ascii")
            mime, _ = mimetypes.guess_type(str(p))
            if not mime:
                mime = f"image/{ext.lstrip('.')}"
            return f"data:{mime};base64,{b64}"

        if ext in {".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz", ".tar.gz", ".tar.bz2"}:
            seven_zip = (Path(__file__).resolve().parent / "libs" / "7z" / "win" / "7za.exe")
            if not seven_zip.exists():
                return "7z 工具不存在: libs/7z/win/7za.exe"

            res = GloscTools.spawn_command(str(seven_zip), ["l", str(p)], encoding="gbk")
            file_list = GloscTools.parse_7z_output(res.output)
            return json.dumps(file_list, ensure_ascii=False)

        # MarkItDown 优先：除图片/压缩包外，先尝试让 MarkItDown 做转换。
        # 若 MarkItDown 无法处理或返回空内容，再回退到原有解析逻辑。
        try:
            md_text = GloscTools._markitdown_convert(p)
            if md_text and md_text.strip():
                return md_text
        except Exception:
            pass

        if ext == ".csv":
            text = GloscTools._decode_text_bytes(data)
            reader = csv.DictReader(text.splitlines())
            rows = list(reader)
            return json.dumps(rows, ensure_ascii=False, indent=2)

        if ext in {".xlsx", ".xls"}:
            as_json = GloscTools._try_read_excel_to_json(p)
            if as_json is not None:
                return as_json
            # Excel 解析失败时，再尝试 MarkItDown（可能依赖可选组件）
            try:
                return GloscTools._markitdown_convert(p)
            except Exception:
                pass

        if ext in {
            ".txt",
            ".md",
            ".json",
            ".js",
            ".ts",
            ".html",
            ".css",
            ".xml",
            ".yaml",
            ".yml",
        }:
            return GloscTools._decode_text_bytes(data)

        return GloscTools._decode_text_bytes(data)

    @staticmethod
    def _decode_text_bytes(data: bytes) -> str:
        for enc in ("utf-8", "utf-8-sig", "gbk", "cp1252"):
            try:
                return data.decode(enc)
            except Exception:
                continue
        return data.decode("utf-8", errors="replace")

    @staticmethod
    def _markitdown_convert(source: Path) -> str:
        md = MarkItDown()
        result = md.convert(source)
        return result.text_content

    @staticmethod
    def _try_read_excel_to_json(source: Path) -> str | None:
        if source.suffix.lower() == ".xlsx":
            try:
                import openpyxl  # type: ignore

                wb = openpyxl.load_workbook(source, data_only=True)
                sheet_names = [s for s in wb.sheetnames if isinstance(s, str) and s.strip()]
                if not sheet_names:
                    return json.dumps([], ensure_ascii=False, indent=2)

                sheets: list[dict[str, Any]] = []
                for name in sheet_names:
                    ws = wb[name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        sheets.append({"sheetName": name, "rows": []})
                        continue

                    header = [str(c).strip() if c is not None else "" for c in rows[0]]
                    normalized_header = [h if h else f"col_{i+1}" for i, h in enumerate(header)]

                    out_rows: list[dict[str, Any]] = []
                    for r in rows[1:]:
                        obj: dict[str, Any] = {}
                        for i, key in enumerate(normalized_header):
                            obj[key] = r[i] if i < len(r) else None
                        out_rows.append(obj)

                    sheets.append({"sheetName": name, "rows": out_rows})

                if len(sheets) == 1:
                    return json.dumps(sheets[0]["rows"], ensure_ascii=False, indent=2)
                return json.dumps({"sheets": sheets}, ensure_ascii=False, indent=2)

            except Exception:
                return None

        if source.suffix.lower() == ".xls":
            try:
                import xlrd  # type: ignore

                book = xlrd.open_workbook(str(source))
                sheet_names = [s for s in book.sheet_names() if isinstance(s, str) and s.strip()]
                if not sheet_names:
                    return json.dumps([], ensure_ascii=False, indent=2)

                sheets: list[dict[str, Any]] = []
                for name in sheet_names:
                    sh = book.sheet_by_name(name)
                    if sh.nrows == 0:
                        sheets.append({"sheetName": name, "rows": []})
                        continue

                    header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
                    normalized_header = [h if h else f"col_{i+1}" for i, h in enumerate(header)]

                    out_rows: list[dict[str, Any]] = []
                    for r in range(1, sh.nrows):
                        obj: dict[str, Any] = {}
                        for c, key in enumerate(normalized_header):
                            v = sh.cell_value(r, c)
                            obj[key] = v
                        out_rows.append(obj)
                    sheets.append({"sheetName": name, "rows": out_rows})

                if len(sheets) == 1:
                    return json.dumps(sheets[0]["rows"], ensure_ascii=False, indent=2)
                return json.dumps({"sheets": sheets}, ensure_ascii=False, indent=2)
            except Exception:
                return None

        return None

    @staticmethod
    def _path_exists(p: Path) -> bool:
        try:
            p.lstat()
            return True
        except FileNotFoundError:
            return False

    @staticmethod
    def _remove_path(target: Path) -> None:
        if target.is_dir():
            shutil.rmtree(target, ignore_errors=True)
        else:
            try:
                target.unlink()
            except FileNotFoundError:
                pass

    @staticmethod
    def _move_path_internal(options: dict[str, Any]) -> dict[str, Any]:
        from_path = str(options.get("from") or "").strip()
        to_path = str(options.get("to") or "").strip()
        if not from_path:
            raise RuntimeError("from 不能为空")
        if not to_path:
            raise RuntimeError("to 不能为空")

        overwrite = bool(options.get("overwrite") or False)
        create_dirs = bool(options.get("createDirs") if options.get("createDirs") is not None else True)

        src = Path(from_path)
        if not src.exists():
            raise RuntimeError("源路径不存在")

        dest = Path(to_path)
        to_looks_like_dir = bool(re.search(r"[\\/]+$", to_path))
        if to_looks_like_dir:
            dest = dest / src.name
        else:
            if dest.exists() and dest.is_dir():
                dest = dest / src.name

        if dest.exists() or dest.is_symlink():
            if not overwrite:
                raise RuntimeError("目标已存在；如需覆盖请设置 overwrite=true")
            GloscTools._remove_path(dest)

        if create_dirs:
            dest.parent.mkdir(parents=True, exist_ok=True)

        try:
            src.rename(dest)
            return {"ok": True, "from": str(src), "to": str(dest)}
        except OSError as e:
            # EXDEV / cross-device
            if getattr(e, "winerror", None) in (17,) or getattr(e, "errno", None) == getattr(os, "EXDEV", 18):
                if src.is_dir():
                    shutil.copytree(src, dest, dirs_exist_ok=False)
                    shutil.rmtree(src)
                else:
                    shutil.copy2(src, dest)
                    src.unlink()
                return {"ok": True, "from": str(src), "to": str(dest)}
            raise

    @staticmethod
    def rename_file(options: dict[str, Any]) -> dict[str, Any]:
        src_path = str(options.get("path") or "").strip()
        new_name = str(options.get("newName") or "").strip()
        overwrite = bool(options.get("overwrite") or False)

        if not src_path:
            raise RuntimeError("path 不能为空")
        if not new_name:
            raise RuntimeError("newName 不能为空")

        base = Path(new_name).name
        if base != new_name:
            raise RuntimeError("newName 只能是文件名，不能包含路径分隔符")

        src = Path(src_path)
        dest = src.parent / new_name
        return GloscTools._move_path_internal(
            {"from": str(src), "to": str(dest), "overwrite": overwrite, "createDirs": True}
        )

    @staticmethod
    def move_file(options: dict[str, Any]) -> dict[str, Any]:
        return GloscTools._move_path_internal(options)

    @staticmethod
    def list_files_recursive(options: dict[str, Any]) -> dict[str, Any]:
        dir_path = str(options.get("dir") or "").strip()
        if not dir_path:
            raise RuntimeError("dir 不能为空")

        limit = int(options.get("limit") or 5000)
        limit = max(1, limit)

        root = Path(dir_path)
        if not root.exists():
            raise RuntimeError("目录不存在")
        if not root.is_dir():
            raise RuntimeError("dir 必须是目录")

        files: list[str] = []
        truncated = False

        for current_root, _, filenames in os.walk(root):
            for name in filenames:
                files.append(str(Path(current_root) / name))
                if len(files) >= limit:
                    truncated = True
                    break
            if truncated:
                break

        return {"ok": True, "dir": dir_path, "files": files, "truncated": truncated}

    @staticmethod
    def write_file(options: dict[str, Any]) -> dict[str, Any]:
        path = options.get("path")
        mode = options.get("mode", "overwrite")
        content = options.get("content", "")
        line = options.get("line")
        old_string = options.get("old_string")
        new_string = options.get("new_string")

        if not path:
            return {"ok": False, "error": "path is required"}

        p = Path(path)

        if mode == "create":
            if p.exists():
                return {"ok": False, "error": "file already exists"}
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(content, encoding="utf-8")
            return {"ok": True, "message": "file created"}

        if not p.exists():
            # return {"ok": False, "error": "file does not exist"}
            # 如果不存在，则创建新文件
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(content, encoding="utf-8")

        if mode == "overwrite":
            p.write_text(content, encoding="utf-8")
            return {"ok": True, "message": "file overwritten"}

        # Read current content
        current_content = p.read_text(encoding="utf-8")
        lines = current_content.splitlines(keepends=True)

        if mode == "insert_line":
            if line is None:
                return {"ok": False, "error": "line is required for insert_line"}
            if line < 1 or line > len(lines) + 1:
                return {"ok": False, "error": "invalid line number"}
            lines.insert(line - 1, content + "\n")
            new_content = "".join(lines)
            p.write_text(new_content, encoding="utf-8")
            return {"ok": True, "message": "line inserted"}

        if mode == "insert_lines":
            if line is None:
                return {"ok": False, "error": "line is required for insert_lines"}
            if line < 1 or line > len(lines) + 1:
                return {"ok": False, "error": "invalid line number"}
            new_lines = content.splitlines(keepends=True)
            for i, nl in enumerate(new_lines):
                lines.insert(line - 1 + i, nl)
            new_content = "".join(lines)
            p.write_text(new_content, encoding="utf-8")
            return {"ok": True, "message": "lines inserted"}

        if mode == "replace":
            if old_string is None or new_string is None:
                return {"ok": False, "error": "old_string and new_string are required for replace"}
            if old_string not in current_content:
                return {"ok": False, "error": "old_string not found"}
            new_content = current_content.replace(old_string, new_string, 1)  # Replace first occurrence
            p.write_text(new_content, encoding="utf-8")
            return {"ok": True, "message": "string replaced"}

        return {"ok": False, "error": "invalid mode"}
